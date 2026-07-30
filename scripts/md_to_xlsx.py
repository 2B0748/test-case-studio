#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
md_to_xlsx.py 鈥?灏?Markdown 娴嬭瘯鐢ㄤ緥琛ㄨ浆鎹负 Excel(.xlsx) 鎴?CSV銆?
鐢ㄦ硶:
    python md_to_xlsx.py <input.md> [--out output.xlsx] [--csv]

璇存槑:
    - 鎸?'## ' 浜岀骇鏍囬鍒囧垎绔犺妭锛屾瘡涓珷鑺傜殑琛ㄦ牸鍐欏叆涓€涓?sheet
      (openpyxl 鍙敤鏃? 鎴栦竴涓嫭绔?CSV锛?-csv 鎴?openpyxl 缂哄け鏃讹級銆?    - 鍚屼竴绔犺妭鍑虹幇澶氬紶琛ㄦ椂锛氳〃澶寸浉鍚屽垯鍚堝苟鏁版嵁琛岋紱琛ㄥご涓嶅悓鍒欎互
      绌鸿 + 鏂拌〃澶撮『搴忚拷鍔犮€?    - 鍗曞厓鏍间腑鐨?<br> / <br/> / <br /> 浼氳浆鎹负 Excel 鍗曞厓鏍煎唴鎹㈣锛?      鏁版嵁琛岃嚜鍔ㄥ紑鍚嚜鍔ㄦ崲琛屼笌椤剁瀵归綈銆?    - 闆朵緷璧栦篃鍙伐浣滐紙鑷姩鍥為€€鍒?CSV锛夛紱瀹夎 openpyxl 鍙幏寰楃湡姝ｇ殑 .xlsx銆?      pip install openpyxl
"""
import argparse
import csv
import os
import re
import sys

TABLE_ROW = re.compile(r"^\s*\|.*\|\s*$")
SEP_CELL = re.compile(r":?-+:?")
BR = re.compile(r"<br\s*/?>", re.IGNORECASE)
BAD_NAME = re.compile(r"[:?*\[\]/\\]")


def split_sections(lines):
    """鎸?'## ' 鏍囬鍒囧垎锛岃繑鍥?[(鏍囬, 琛屽垪琛?, ...]"""
    sections = []
    current_title = "鐢ㄤ緥"
    current_lines = []
    for line in lines:
        m = re.match(r"^##\s+(.*)$", line)
        if m:
            if current_lines:
                sections.append((current_title, current_lines))
            current_title = m.group(1).strip()
            current_lines = []
        else:
            current_lines.append(line)
    if current_lines:
        sections.append((current_title, current_lines))
    return sections


def parse_tables(block_lines):
    """瑙ｆ瀽涓€缁勮涓殑鎵€鏈?Markdown 琛ㄦ牸锛岃繑鍥?[(header, rows), ...]銆?    琛ㄦ牸涔嬮棿浠ラ潪琛ㄦ牸琛岋紙绌鸿銆佽鏄庢枃瀛楃瓑锛夊垎闅斻€?""
    tables = []
    cur = []

    def flush():
        if cur:
            data = [r for r in cur if not is_sep_row(r)]
            if data:
                tables.append((data[0], data[1:]))
            cur.clear()

    for line in block_lines:
        if TABLE_ROW.match(line):
            cells = [clean_cell(c) for c in line.strip().strip("|").split("|")]
            cur.append(cells)
        else:
            flush()
    flush()
    return tables


def is_sep_row(row):
    return all(SEP_CELL.fullmatch(c) for c in row if c != "")


def clean_cell(text):
    """鍗曞厓鏍兼竻娲楋細<br> 绯诲垪鏍囩杞崲涓烘崲琛岀銆?""
    return BR.sub("\n", text.strip())


def safe_sheet(name, taken):
    """Excel sheet 鍚嶉檺鍒讹細鈮?1 瀛楃锛岀鐢?[:?*/\\]锛涢噸鍚嶆椂鑷姩杩藉姞搴忓彿銆?""
    base = BAD_NAME.sub("_", name).strip() or "Sheet"
    base = base[:31]
    candidate, i = base, 2
    while candidate in taken:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    taken.add(candidate)
    return candidate


def safe_filename(name):
    return BAD_NAME.sub("_", name).strip()[:31] or "sheet"


def pad_row(row, width):
    if len(row) < width:
        return row + [""] * (width - len(row))
    return row[:width]


def to_xlsx(sections, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[warn] 鏈娴嬪埌 openpyxl锛岃嚜鍔ㄥ洖閫€涓?CSV 杈撳嚭銆?)
        print("       瀹夎: pip install openpyxl  浠ヨ幏寰楃湡姝ｇ殑 .xlsx銆?)
        base = os.path.splitext(out_path)[0]
        to_csv_fallback(sections, base)
        return

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)
    taken = set()

    for title, lines in sections:
        tables = parse_tables(lines)
        if not tables:
            continue
        ws = wb.create_sheet(title=safe_sheet(title, taken))
        max_cols = 0
        prev_header = None
        for header, body in tables:
            max_cols = max(max_cols, len(header))
            same = header == prev_header
            if prev_header is not None and not same:
                ws.append([])  # 涓嶅悓琛ㄥご涔嬮棿绌轰竴琛?            if not same:
                ws.append(header)
                for c in ws[ws.max_row]:
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = header_align
            for r in body:
                ws.append(pad_row(r, len(header)))
                for c in ws[ws.max_row]:
                    c.alignment = data_align
            prev_header = header
        # 鍒楀鑷€傚簲锛堜腑鏂囨寜 2 璁★紝澶氳鍗曞厓鏍兼寜鏈€闀夸竴琛屼及绠楋級
        for col_idx in range(1, max_cols + 1):
            char_w = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                v = row[0].value
                if v is None:
                    continue
                w = max(
                    (sum(2 if ord(ch) > 0x2E80 else 1 for ch in seg)
                     for seg in str(v).split("\n")),
                    default=0,
                )
                char_w = max(char_w, w)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(char_w + 2, 10), 60)
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        print("[warn] 鏈В鏋愬埌浠讳綍 Markdown 琛ㄦ牸锛屾湭鐢熸垚鏂囦欢銆?)
        sys.exit(1)
    wb.save(out_path)
    print(f"[ok] 宸茬敓鎴?Excel: {out_path}  (sheet 鏁? {len(wb.sheetnames)})")


def to_csv_fallback(sections, base):
    made = 0
    for i, (title, lines) in enumerate(sections):
        tables = parse_tables(lines)
        if not tables:
            continue
        fname = f"{base}_{i}_{safe_filename(title)}.csv"
        with open(fname, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            prev_header = None
            for header, body in tables:
                same = header == prev_header
                if prev_header is not None and not same:
                    w.writerow([])
                if not same:
                    w.writerow(header)
                for r in body:
                    w.writerow(pad_row(r, len(header)))
                prev_header = header
        print(f"[ok] 宸茬敓鎴?CSV: {fname}")
        made += 1
    if made == 0:
        print("[warn] 鏈В鏋愬埌浠讳綍 Markdown 琛ㄦ牸銆?)
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser(description="Markdown 鐢ㄤ緥琛?鈫?Excel/CSV")
    ap.add_argument("input", help="杈撳叆鐨?Markdown 鏂囦欢")
    ap.add_argument("--out", default=None, help="杈撳嚭 .xlsx 璺緞")
    ap.add_argument("--csv", action="store_true", help="寮哄埗杈撳嚭 CSV")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        print(f"[error] 鏂囦欢涓嶅瓨鍦? {args.input}")
        sys.exit(1)

    with open(args.input, encoding="utf-8") as f:
        lines = f.readlines()

    sections = split_sections(lines)
    if args.csv:
        base = args.out or os.path.splitext(args.input)[0]
        to_csv_fallback(sections, base)
    else:
        out = args.out or (os.path.splitext(args.input)[0] + ".xlsx")
        to_xlsx(sections, out)


if __name__ == "__main__":
    main()
